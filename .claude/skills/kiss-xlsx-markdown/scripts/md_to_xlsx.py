#!/usr/bin/env python3
"""Rebuild an .xlsx from the Markdown produced by xlsx_to_md.py.

Usage:
    python md_to_xlsx.py input.md output.xlsx
        [--meta input.meta.json]
        [--template original.xlsx]
        [--recalc]
"""
from __future__ import annotations

import sys
from pathlib import Path

# Run the venv bootstrap BEFORE any third-party imports. If ./.venv exists in
# the user's cwd, this re-execs the script under that interpreter; if not, it
# prints setup instructions unless --system-python was passed.
if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).parent))
    from venv_bootstrap import ensure_env  # noqa: E402
    ensure_env(__file__, skill_name="kiss-xlsx-markdown")

import argparse  # noqa: E402
import json  # noqa: E402
import re  # noqa: E402
import shutil  # noqa: E402
import subprocess  # noqa: E402


def _pip_hint(packages: str) -> str:
    """Return a pip install hint that works on Windows, macOS, and Linux."""
    if sys.platform.startswith("win"):
        return f"pip install {packages}"
    return f"pip install {packages} --break-system-packages"


try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    sys.exit(f"Install openpyxl: {_pip_hint('openpyxl pyyaml')}")

try:
    import yaml
except ImportError:
    sys.exit(f"Install pyyaml: {_pip_hint('pyyaml')}")


SHEET_START_RE = re.compile(r"<!--\s*sheet:start\s+(?P<attrs>[^>]+?)-->")
SHEET_END_RE = re.compile(r"<!--\s*sheet:end\s*-->")
META_FENCE_RE = re.compile(r"```yaml\s+meta\s*\n(?P<yaml>.*?)\n```", re.DOTALL)


def parse_attr_block(attr_str: str) -> dict:
    out: dict = {}
    for m in re.finditer(r'(\w+)="([^"]*)"', attr_str):
        out[m.group(1)] = m.group(2)
    return out


def parse_frontmatter(text: str) -> tuple[dict, str]:
    if not text.startswith("---"):
        return {}, text
    end = text.find("\n---", 3)
    if end == -1:
        return {}, text
    fm_text = text[3:end].strip()
    body = text[end + 4 :].lstrip("\n")
    return yaml.safe_load(fm_text) or {}, body


def split_sheets(body: str) -> list[dict]:
    sheets: list[dict] = []
    pos = 0
    while True:
        m = SHEET_START_RE.search(body, pos)
        if not m:
            break
        end_m = SHEET_END_RE.search(body, m.end())
        if not end_m:
            raise ValueError("Unclosed <!-- sheet:start --> block")
        attrs = parse_attr_block(m.group("attrs"))
        content = body[m.end() : end_m.start()]
        sheet_meta = {}
        fm = META_FENCE_RE.search(content)
        if fm:
            sheet_meta = yaml.safe_load(fm.group("yaml")) or {}
            content = content[: fm.start()] + content[fm.end() :]
        sheets.append(
            {
                "name": attrs.get("name", f"Sheet{len(sheets) + 1}"),
                "id": attrs.get("id", str(len(sheets) + 1)),
                "hidden": attrs.get("hidden", "false"),
                "meta": sheet_meta,
                "content": content.strip(),
            }
        )
        pos = end_m.end()
    return sheets


def parse_table(content: str) -> list[list[str]]:
    """Return raw rows as lists of strings, preserving formula backticks."""
    rows: list[list[str]] = []
    caption_attrs: dict = {}
    for line in content.splitlines():
        s = line.strip()
        if s.startswith("|") and s.endswith("|"):
            inner = s[1:-1]
            # header separator?
            if re.match(r"^\s*[:\- |]+\s*$", inner):
                continue
            cells = [c.strip() for c in _split_pipes(inner)]
            rows.append(cells)
        elif s.startswith(": {") and s.endswith("}"):
            # caption/attr block
            body = s[3:-1]
            for tok in body.split():
                if "=" in tok:
                    k, v = tok.split("=", 1)
                    caption_attrs[k.strip()] = v.strip('"')
    # The markdown header row corresponds to the first row of the sheet's
    # used range (see xlsx_to_md.sheet_to_markdown). Keep it as data unless
    # caption attribute explicitly opts out.
    if caption_attrs.get("no_header") == "true" and rows:
        rows = rows[1:]
    start_row = int(caption_attrs.get("start_row", 1))
    return rows, start_row


def _split_pipes(s: str) -> list[str]:
    # Handle escaped pipes \|
    out = []
    buf = []
    i = 0
    while i < len(s):
        ch = s[i]
        if ch == "\\" and i + 1 < len(s) and s[i + 1] == "|":
            buf.append("|")
            i += 2
            continue
        if ch == "|":
            out.append("".join(buf))
            buf = []
            i += 1
            continue
        buf.append(ch)
        i += 1
    out.append("".join(buf))
    return out


def unescape_cell(val: str) -> object:
    v = val.strip()
    if not v:
        return None
    # Formula: backtick-wrapped starting with `=`
    if v.startswith("`") and v.endswith("`") and v[1:2] == "=":
        return v[1:-1]
    # Raw formula without backticks (tolerant)
    if v.startswith("="):
        return v
    if v.upper() == "TRUE":
        return True
    if v.upper() == "FALSE":
        return False
    # Try numeric parse — strip thousands commas, percent, currency.
    vnum = v.replace(",", "")
    if vnum.startswith("$"):
        vnum = vnum[1:]
    percent = vnum.endswith("%")
    if percent:
        vnum = vnum[:-1]
    try:
        n = float(vnum)
        if percent:
            n = n / 100.0
        if n.is_integer() and "%" not in v and "." not in v:
            return int(n)
        return n
    except ValueError:
        pass
    # Date (ISO)
    if re.match(r"^\d{4}-\d{2}-\d{2}(T\d{2}:\d{2}:\d{2})?", v):
        import datetime as _dt
        try:
            if "T" in v:
                return _dt.datetime.fromisoformat(v)
            return _dt.date.fromisoformat(v)
        except Exception:
            pass
    return v.replace("<br>", "\n")


def apply_sheet_meta(ws, sheet_meta: dict) -> None:
    if not sheet_meta:
        return
    if sheet_meta.get("freeze_panes"):
        ws.freeze_panes = sheet_meta["freeze_panes"]
    if sheet_meta.get("tab_color"):
        ws.sheet_properties.tabColor = sheet_meta["tab_color"]
    if sheet_meta.get("auto_filter"):
        ws.auto_filter.ref = sheet_meta["auto_filter"]
    for rng in sheet_meta.get("merged_cells", []) or []:
        ws.merge_cells(rng)
    for letter, width in (sheet_meta.get("column_widths") or {}).items():
        ws.column_dimensions[str(letter)].width = float(width)
    for row, height in (sheet_meta.get("row_heights") or {}).items():
        ws.row_dimensions[int(row)].height = float(height)


def apply_cell_styles(ws, cells_style: dict) -> None:
    for coord, style in (cells_style or {}).items():
        cell = ws[coord]
        if "number_format" in style:
            cell.number_format = style["number_format"]
        if "font" in style:
            f = style["font"]
            cell.font = Font(
                name=f.get("name"), size=f.get("size") or 11,
                bold=f.get("bold", False), italic=f.get("italic", False),
                color=f.get("color") or "FF000000",
            )
        if "fill" in style and style["fill"].get("patternType"):
            fill = style["fill"]
            cell.fill = PatternFill(
                patternType=fill.get("patternType"),
                fgColor=fill.get("fgColor") or "FFFFFFFF",
                bgColor=fill.get("bgColor") or "FFFFFFFF",
            )
        if "border" in style:
            b = style["border"]
            cell.border = Border(
                top=Side(style=b.get("top")),
                bottom=Side(style=b.get("bottom")),
                left=Side(style=b.get("left")),
                right=Side(style=b.get("right")),
            )
        if "alignment" in style:
            a = style["alignment"]
            cell.alignment = Alignment(
                horizontal=a.get("horizontal"),
                vertical=a.get("vertical"),
                wrap_text=a.get("wrap_text", False),
            )


def rebuild(md_path: Path, out_path: Path, template: Path | None, meta_path: Path | None,
            recalc: bool) -> None:
    text = md_path.read_text(encoding="utf-8")
    fm, body = parse_frontmatter(text)

    meta: dict = {}
    if meta_path is None and fm.get("xlsx_meta"):
        meta_path = md_path.parent / fm["xlsx_meta"]
    if meta_path and Path(meta_path).exists():
        meta = json.loads(Path(meta_path).read_text(encoding="utf-8"))

    if template and template.exists():
        wb = load_workbook(template, data_only=False)
        # Drop existing sheets
        for name in list(wb.sheetnames):
            del wb[name]
    else:
        wb = Workbook()
        # Remove the default sheet; we'll add our own.
        wb.remove(wb.active)

    sheet_blocks = split_sheets(body)
    sheets_order = fm.get("sheets_order") or [s["name"] for s in sheet_blocks]
    sheet_by_name = {s["name"]: s for s in sheet_blocks}

    for name in sheets_order:
        if name not in sheet_by_name:
            continue
        block = sheet_by_name[name]
        ws = wb.create_sheet(title=name)
        if block["hidden"] == "true":
            ws.sheet_state = "hidden"
        elif block["hidden"] == "veryhidden":
            ws.sheet_state = "veryHidden"

        rows, start_row = parse_table(block["content"])
        # We dropped the header row during parse; the first data row goes to Excel row 1
        # unless the user specified otherwise.
        for r_idx, row in enumerate(rows, start=start_row):
            for c_idx, cell_val in enumerate(row, start=1):
                val = unescape_cell(cell_val)
                if val is None:
                    continue
                ws.cell(row=r_idx, column=c_idx, value=val)

        apply_sheet_meta(ws, block["meta"])
        sheet_meta_sidecar = (meta.get("sheets") or {}).get(name, {})
        # Sidecar can supply any key the Markdown meta block didn't.
        combined = {**sheet_meta_sidecar, **block["meta"]}
        apply_sheet_meta(ws, combined)
        apply_cell_styles(ws, sheet_meta_sidecar.get("cells") or {})

    # Workbook-scoped defined names.
    for nm, value in (fm.get("defined_names") or {}).items():
        try:
            from openpyxl.workbook.defined_name import DefinedName
            wb.defined_names[nm] = DefinedName(name=nm, attr_text=value)
        except Exception:
            pass
    for nm, value in (meta.get("defined_names") or {}).items():
        if nm in (fm.get("defined_names") or {}):
            continue
        try:
            from openpyxl.workbook.defined_name import DefinedName
            wb.defined_names[nm] = DefinedName(name=nm, attr_text=value)
        except Exception:
            pass

    if fm.get("title"):
        wb.properties.title = fm["title"]
    if fm.get("author"):
        wb.properties.creator = fm["author"]

    wb.save(out_path)
    print(f"Wrote {out_path}")

    if recalc:
        here = Path(__file__).parent
        recalc_script = here / "recalc.py"
        if recalc_script.exists():
            subprocess.run([sys.executable, str(recalc_script), str(out_path)], check=False)


def main() -> None:
    ap = argparse.ArgumentParser(description="Rebuild .xlsx from Markdown.")
    ap.add_argument("input", type=Path)
    ap.add_argument("output", type=Path)
    ap.add_argument("--meta", type=Path, default=None)
    ap.add_argument("--template", type=Path, default=None)
    ap.add_argument("--recalc", action="store_true")
    args = ap.parse_args()
    rebuild(args.input, args.output, args.template, args.meta, args.recalc)


if __name__ == "__main__":
    main()
