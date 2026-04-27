#!/usr/bin/env python3
"""Recalculate formulas in an .xlsx via LibreOffice (headless).

Minimal implementation: opens and re-saves the workbook through soffice, then
scans for Excel error strings in the cached values. For a full-featured
implementation, use the recalc.py shipped with the base xlsx skill.
"""
from __future__ import annotations

import sys
from pathlib import Path

# Run the venv bootstrap BEFORE any third-party imports. If ./.venv exists in
# the user's cwd, this re-execs the script under that interpreter; if not, it
# prints setup instructions unless --system-python was passed.
if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).parent))
    from venv_bootstrap import ensure_env  # noqa: E402
    ensure_env(__file__, skill_name="kiss-xlsx-markdown")

import argparse  # noqa: E402
import json  # noqa: E402
import shutil  # noqa: E402
import subprocess  # noqa: E402


def _pip_hint(packages: str) -> str:
    """Return a pip install hint that works on Windows, macOS, and Linux."""
    if sys.platform.startswith("win"):
        return f"pip install {packages}"
    return f"pip install {packages} --break-system-packages"


try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit(f"Install openpyxl: {_pip_hint('openpyxl')}")


ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")


def find_soffice() -> str | None:
    """Locate LibreOffice's ``soffice`` executable across macOS, Linux, and Windows.

    Falls back to common install locations not always present on PATH
    (the macOS ``.app`` bundle, standard Windows install dirs, snap).
    """
    found = shutil.which("soffice") or shutil.which("libreoffice")
    if found:
        return found
    candidates: list[str] = []
    if sys.platform == "darwin":
        candidates += [
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        ]
    elif sys.platform.startswith("win"):
        candidates += [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
    else:
        candidates += [
            "/usr/bin/soffice",
            "/usr/local/bin/soffice",
            "/snap/bin/libreoffice",
        ]
    for c in candidates:
        if Path(c).exists():
            return c
    return None


def recalc(path: Path, timeout: int) -> dict:
    soffice = find_soffice()
    if not soffice:
        return {"status": "skipped", "reason": "LibreOffice not found"}
    outdir = path.parent
    result = subprocess.run(
        [soffice, "--headless", "--calc", "--convert-to", "xlsx",
         "--outdir", str(outdir), str(path)],
        capture_output=True, text=True, timeout=timeout,
    )
    if result.returncode != 0:
        return {"status": "error", "stderr": result.stderr}

    wb = load_workbook(path, data_only=True)
    errors_by_kind: dict[str, dict] = {}
    total = 0
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v in ERRORS:
                    total += 1
                    bucket = errors_by_kind.setdefault(v, {"count": 0, "locations": []})
                    bucket["count"] += 1
                    bucket["locations"].append(f"{sheet}!{cell.coordinate}")
    if total == 0:
        return {"status": "success", "total_errors": 0}
    return {"status": "errors_found", "total_errors": total, "error_summary": errors_by_kind}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("path", type=Path)
    ap.add_argument("timeout", type=int, nargs="?", default=60)
    args = ap.parse_args()
    print(json.dumps(recalc(args.path, args.timeout), indent=2))


if __name__ == "__main__":
    main()
