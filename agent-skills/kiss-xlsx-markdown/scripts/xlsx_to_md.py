#!/usr/bin/env python3
"""Export an .xlsx to Markdown + assets + meta.json sidecar.

Usage:
    python xlsx_to_md.py input.xlsx output.md [--assets-dir DIR]
                                               [--no-meta]
                                               [--values-only]

See ../references/format-spec.md for the dialect emitted.
"""
from __future__ import annotations

import sys
from pathlib import Path

# Run the venv bootstrap BEFORE any third-party imports. If ./.venv exists in
# the user's cwd, this re-execs the script under that interpreter; if not, it
# prints setup instructions unless --system-python was passed.
if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).parent))
    from _venv_bootstrap import ensure_env  # noqa: E402
    ensure_env(__file__, skill_name="kiss-xlsx-markdown")

import argparse  # noqa: E402
import json  # noqa: E402


def _pip_hint(packages: str) -> str:
    """Return a pip install hint that works on Windows, macOS, and Linux."""
    if sys.platform.startswith("win"):
        return f"pip install {packages}"
    return f"pip install {packages} --break-system-packages"


try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import TYPE_FORMULA
except ImportError:
    sys.exit(f"Install openpyxl: {_pip_hint('openpyxl pyyaml')}")

try:
    import yaml
except ImportError:
    sys.exit(f"Install pyyaml: {_pip_hint('pyyaml')}")


def fmt_cell(cell, values_only: bool, cached_wb) -> str:
    """Render a single cell as a Markdown table cell body."""
    if cell.value is None:
        return ""
    if cell.data_type == TYPE_FORMULA and not values_only:
        return f"`{cell.value}`"
    # Cached or values-only
    val = cell.value
    if cell.data_type == TYPE_FORMULA and values_only and cached_wb is not None:
        try:
            cached = cached_wb[cell.parent.title][cell.coordinate].value
            val = cached
        except Exception:
            pass
    if val is None:
        return ""
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, (int, float)):
        # Apply a loose rendering using the number format.
        nf = cell.number_format or "General"
        return render_number(val, nf)
    # datetime
    try:
        import datetime as _dt
        if isinstance(val, (_dt.datetime, _dt.date)):
            return val.isoformat()
    except Exception:
        pass
    s = str(val)
    s = s.replace("|", "\\|").replace("\n", "<br>")
    return s


def render_number(val: float, fmt: str) -> str:
    """Very small number-format renderer. Full fidelity lives in meta.json."""
    try:
        if "%" in fmt:
            # percentage
            digits = fmt.count("0") - 1  # count after decimal
            if "." in fmt:
                after = fmt.split(".", 1)[1]
                digits = after.count("0")
            else:
                digits = 0
            return f"{val * 100:.{digits}f}%"
        if "$" in fmt or "¤" in fmt:
            return f"${val:,.0f}" if "." not in fmt else f"${val:,.2f}"
        if "," in fmt:
            if isinstance(val, float) and not val.is_integer():
                return f"{val:,.2f}"
            return f"{int(val):,}"
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val)
    except Exception:
        return str(val)


def _rgb(color_obj) -> str | None:
    """openpyxl Color can be rgb, theme, indexed, or typed None. Return only
    a valid 8-char aRGB hex string so the sidecar is safe to round-trip.
    Theme/indexed colors are skipped — rebuilt sheets will inherit template.
    """
    if color_obj is None:
        return None
    # Skip non-rgb color kinds (theme, indexed, tint-only).
    ctype = getattr(color_obj, "type", None)
    if ctype and ctype != "rgb":
        return None
    try:
        val = color_obj.rgb if hasattr(color_obj, "rgb") else None
    except Exception:
        return None
    if not isinstance(val, str):
        return None
    # Accept 8-char aRGB hex only (Excel requires this shape on rebuild).
    if len(val) == 8 and all(ch in "0123456789abcdefABCDEF" for ch in val):
        return val.upper()
    if len(val) == 6 and all(ch in "0123456789abcdefABCDEF" for ch in val):
        return "FF" + val.upper()
    return None


def extract_sheet_meta(ws) -> dict:
    merged = [str(r) for r in ws.merged_cells.ranges]
    col_widths = {}
    for letter, dim in ws.column_dimensions.items():
        if dim.width:
            col_widths[letter] = round(dim.width, 2)
    row_heights = {}
    for row, dim in ws.row_dimensions.items():
        if dim.height:
            row_heights[int(row)] = round(dim.height, 2)
    meta = {
        "id": ws.sheet_properties.tabColor.value if False else None,
        "tab_color": _rgb(ws.sheet_properties.tabColor),
        "freeze_panes": ws.freeze_panes or None,
        "auto_filter": ws.auto_filter.ref if ws.auto_filter and ws.auto_filter.ref else None,
        "merged_cells": merged,
        "column_widths": dict(sorted(col_widths.items())),
        "row_heights": dict(sorted(row_heights.items())),
    }
    return {k: v for k, v in meta.items() if v not in (None, [], {}, "")}


def sheet_bounds(ws):
    return ws.min_row or 1, ws.max_row or 1, ws.min_column or 1, ws.max_column or 1


def sheet_to_markdown(ws, values_only: bool, cached_wb) -> str:
    min_r, max_r, min_c, max_c = sheet_bounds(ws)
    if max_r < min_r or max_c < min_c:
        return "_(empty sheet)_\n"
    # Header row = first row of the used range
    rows = []
    for r in range(min_r, max_r + 1):
        row_cells = []
        for c in range(min_c, max_c + 1):
            cell = ws.cell(row=r, column=c)
            row_cells.append(fmt_cell(cell, values_only, cached_wb))
        rows.append(row_cells)
    # Column letters for the header row labels
    col_letters = [get_column_letter(c) for c in range(min_c, max_c + 1)]
    # Emit with *literal* column letters as header only if first row is non-textual.
    header = rows[0]
    data_rows = rows[1:]
    if not any(header):
        header = col_letters
    lines = []
    lines.append("| " + " | ".join(h or "" for h in header) + " |")
    lines.append("|" + "|".join([" --- "] * len(header)) + "|")
    for row in data_rows:
        # Pad row to header width
        while len(row) < len(header):
            row.append("")
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines) + "\n"


def extract_cell_styles(ws) -> dict:
    cells = {}
    min_r, max_r, min_c, max_c = sheet_bounds(ws)
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None and not cell.has_style:
                continue
            style: dict = {}
            if cell.number_format and cell.number_format != "General":
                style["number_format"] = cell.number_format
            font = cell.font
            if font and (font.name or font.bold or font.italic or font.color):
                style["font"] = {
                    "name": font.name,
                    "size": float(font.size) if font.size else None,
                    "bold": bool(font.bold),
                    "italic": bool(font.italic),
                    "color": _rgb(font.color),
                }
            fill = cell.fill
            if fill and fill.patternType and fill.patternType != "none":
                style["fill"] = {
                    "patternType": fill.patternType,
                    "fgColor": _rgb(fill.fgColor),
                    "bgColor": _rgb(fill.bgColor),
                }
            border = cell.border
            if border and any(getattr(getattr(border, s), "style", None) for s in ("top", "bottom", "left", "right")):
                style["border"] = {
                    "top": border.top.style,
                    "bottom": border.bottom.style,
                    "left": border.left.style,
                    "right": border.right.style,
                }
            al = cell.alignment
            if al and (al.horizontal or al.vertical or al.wrap_text):
                style["alignment"] = {
                    "horizontal": al.horizontal,
                    "vertical": al.vertical,
                    "wrap_text": bool(al.wrap_text),
                }
            if style:
                cells[cell.coordinate] = style
    return cells


def workbook_to_markdown(wb, dest: Path, assets_dir: Path, values_only: bool,
                         cached_wb, write_meta: bool) -> None:
    meta = {
        "core_properties": {
            "title": wb.properties.title or "",
            "author": wb.properties.creator or "",
            "created": wb.properties.created.isoformat() if wb.properties.created else "",
            "modified": wb.properties.modified.isoformat() if wb.properties.modified else "",
        },
        "defined_names": {},
        "sheets": {},
    }
    for name in wb.defined_names:
        try:
            defn = wb.defined_names[name]
            if hasattr(defn, "attr_text"):
                meta["defined_names"][name] = defn.attr_text
        except Exception:
            pass

    fm = {
        "title": wb.properties.title or "",
        "author": wb.properties.creator or "",
        "created": wb.properties.created.isoformat() if wb.properties.created else "",
        "modified": wb.properties.modified.isoformat() if wb.properties.modified else "",
        "sheets_order": wb.sheetnames,
    }
    if write_meta:
        fm["xlsx_meta"] = f"{dest.stem}.meta.json"

    body_parts = []
    for idx, name in enumerate(wb.sheetnames, start=1):
        ws = wb[name]
        sheet_meta = extract_sheet_meta(ws)
        if write_meta:
            cells = extract_cell_styles(ws)
            meta["sheets"][name] = {**sheet_meta, "id": idx, "cells": cells}

        body_parts.append(f'<!-- sheet:start name="{name}" id="{idx}" -->')
        meta_block = {k: v for k, v in sheet_meta.items() if k in (
            "tab_color", "freeze_panes", "auto_filter", "merged_cells",
            "column_widths", "row_heights"
        )}
        if meta_block:
            body_parts.append("```yaml meta")
            body_parts.append(yaml.safe_dump(meta_block, sort_keys=True).rstrip())
            body_parts.append("```")
        body_parts.append("")
        body_parts.append(sheet_to_markdown(ws, values_only, cached_wb))
        body_parts.append("<!-- sheet:end -->\n")

    header = "---\n" + yaml.safe_dump(fm, sort_keys=True, default_flow_style=False).rstrip() + "\n---\n"
    text = header + "\n" + "\n".join(body_parts)
    text = "\n".join(ln.rstrip() for ln in text.splitlines()) + "\n"
    dest.write_text(text, encoding="utf-8")

    if write_meta:
        (dest.parent / f"{dest.stem}.meta.json").write_text(
            json.dumps(meta, indent=2, sort_keys=True), encoding="utf-8"
        )

    print(f"Wrote {dest}")
    print(f"  assets -> {assets_dir}")
    if write_meta:
        print(f"  meta   -> {dest.parent / (dest.stem + '.meta.json')}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Convert .xlsx to extended Markdown.")
    ap.add_argument("input", type=Path)
    ap.add_argument("output", type=Path)
    ap.add_argument("--assets-dir", type=Path, default=None)
    ap.add_argument("--no-meta", action="store_true")
    ap.add_argument("--values-only", action="store_true")
    args = ap.parse_args()

    wb = load_workbook(args.input, data_only=False)
    cached_wb = load_workbook(args.input, data_only=True) if args.values_only else None

    assets_dir = args.assets_dir or args.output.parent / f"{args.output.stem}.assets"
    assets_dir.mkdir(parents=True, exist_ok=True)

    workbook_to_markdown(wb, args.output, assets_dir, args.values_only, cached_wb,
                         write_meta=not args.no_meta)


if __name__ == "__main__":
    main()
